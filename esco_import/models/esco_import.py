from odoo import fields, models, api, _
from datetime import datetime
import openpyxl
from odoo.exceptions import UserError, ValidationError, AccessError
from io import BytesIO
import tempfile
import base64
import os


class SalesImport(models.Model):
    _name = 'sale.order.imports'
    _description = 'This module is used to upload xls files and create a sale order query from it'

    name = fields.Char(string='Sequence Number', default=lambda self: _('New'), required=True, copy=False)
    file_upload = fields.Binary('File', required=True)
    file_path = fields.Char('File Path')
    sale_order_ids = fields.One2many('sale.order', 'sales_import_id', 'Sales ID')
    sale_order_ids_count = fields.Integer('SO Count')
    # Fields specifying custom line logic
    state = fields.Selection(
        selection=[
            ('new', "New"),
            ('done', "Done"),
        ],
        default='new')

    def unlink(self):
        if self.state == 'done':
            raise ValidationError('You can not delete a record in done state.')
        elif self.state == 'new':
            return super(SalesImport, self).unlink()

    def check_if_done_before(self):
        flag = False
        sale_check = self.env['sale.order.imports'].search([('file_upload', '=', self.file_upload), ('state', '=', 'done')])
        if sale_check:
            flag = True
        elif not sale_check:
            flag = False

        return flag
    def confirm_sales(self):
        sale_ids = self.env['sale.order'].search([('sales_import_id', '=', self.id)])
        for rec in sale_ids:
            rec.action_confirm()

        #set status to done
        self.set_to_done()

    def compute_count(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Sale Orders',
            'res_model': 'sale.order',
            'view_type': 'form',
            'view_mode': 'tree,form',
            'domain': [('sales_import_id', '=', self.id)],
            'target': 'current',
        }

    def set_to_done(self):
        self.state = 'done'

    @api.model_create_multi
    def create(self, vals_list):
        res = []
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('sale.order.imports') or _('New')
            res = super(SalesImport, self).create(vals_list)
        return res

    def get_xls_values(self):

        # Load the workbook
        # wb = openpyxl.load_workbook(self.file_path)
        wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file_upload)))

        # Select the active worksheet
        ws = wb.active

        # Access the values of cells in the worksheet
        count = 0
        flag = False
        sale_order_heeader = []
        sale_order_line = []
        sale_order_count = 0
        sale_order_id = 0
        list_of_sale_order_ids = []

        for row in ws.iter_rows():
            count += 1
            order_lines = count + 5
            so = []

            #check if new order
            if ws.cell(count, 26).value == 'رقم الفترة المخصصة':
                #set flag to true
                flag = True
                # get customer and date details
                rows = count + 2
                date_object = datetime.strptime(ws.cell(count, 26 - 11).value, '%d.%m.%Y')
                new_date_string = date_object.strftime('%Y-%m-%d')
                customer_id = self.env['res.partner'].search([('name', '=', ws.cell(count + 2, 26 - 7).value)])

                sale_order_heeader.append({
                    'customer_name': customer_id.id,
                    'sale_order_date': new_date_string,
                    'sale_person': ws.cell(count + 1, 26 - 7).value,
                    'origin': ws.cell(count, 26 - 6).value,
                })

            # check if sale order hold lines
            if flag == True and ws.cell(order_lines, 26).value:
                # get order lines
                if ws.cell(order_lines, 26).value == ':خصم نوع الدفع %':
                    flag = False

                elif flag and ws.cell(order_lines, 26 - 7).value != '© برنامج مراقبة المبيعات ETI 2023':
                    sale_order_line.append({
                        'item_code': ws.cell(order_lines, 26).value,
                        'item_name': ws.cell(order_lines, 26 - 7).value,
                        'item_qty': ws.cell(order_lines, 26 - 9).value,
                        'item_uom': ws.cell(order_lines, 26 - 11).value,
                        'item_price': ws.cell(order_lines, 26 - 15).value,
                    })
                # print("Taxes :", ws.cell(order_lines, 26 - 18).value)

            #check if end of order
            if ws.cell(count, 26-1).value == ':خصم نوع الدفع %':
                #set flag to false
                flag = False
                sale_order_count += 1
                result = self.check_if_done_before()
                if result:
                    raise ValidationError('You created a query with the same path file before, are you sure to complete!.')
                    break

                #create sale order details
                for rec in sale_order_heeader:
                    so = self.env['sale.order'].create({
                        'partner_id': rec['customer_name'],
                        'date_order': rec['sale_order_date'],
                        'user_id': self.env['res.users'].search([('name', '=', rec['sale_person'])]).id,
                        'origin': rec['origin'],
                        'state': 'draft',
                        'sales_import_id': self.id,
                    })
                    sale_order_id = so

                for row in sale_order_line:
                    print("origin :", rec['sale_order_date'])
                    uom = self.env['uom.uom'].search([('name', '=', row['item_uom'])])
                    product_templ = self.env['product.template'].search([('default_code', '=', row['item_code'])])
                    
                    print("SO ID :", sale_order_id.id)
                    print("Itme name :", row['item_name'])
                    print("product id :", product_templ.id)
                    print("product_uom_qty :", row['item_qty'])
                    print("product_uom :", uom.id)
                    print("price_unit :", row['item_price'])

                    sol_new = so.write({
                        'order_line': [
                            (0, 0, {
                                'order_id': sale_order_id.id,
                                'name': row['item_name'],
                                'product_id': product_templ.id,
                                'product_uom_qty': row['item_qty'],
                                'product_uom': uom.id,
                                'price_unit': row['item_price'],
                                'display_type': False,
                            })
                        ]
                    })
                list_of_sale_order_ids.append(sale_order_id.id)



                #empty the sale order and sale order line for new one
                sale_order_heeader = []
                sale_order_line = []


        print("Sale Orders Count", sale_order_count)
        #count number of sale orders created
        self.sale_order_ids_count = self.env['sale.order'].search_count([('sales_import_id', '=', self.id)])

        #call the sales confirmation method to confirm and create stock pick and invoice for sale order and chane current module state to done!
        self.confirm_sales()

class SalesOrder(models.Model):
    _inherit = 'sale.order'

    sales_import_id = fields.Many2one('sale.order.imports', 'Sales Import ID')
